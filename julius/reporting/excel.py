"""Planilha do relatório, para quem precisa filtrar, ordenar e recortar.

O HTML é para ler; a planilha é para trabalhar em cima. Uma aba por bucket de
prioridade, mais saúde da coleta e premissas — que é onde a discussão sobre um
número normalmente termina.

**Nada é calculado aqui.** As mesmas view models que o HTML consome alimentam a
planilha, então os dois não têm como divergir. Se um número está errado, está
errado nos dois, e o conserto é um só.

Valor monetário vai como número, não como texto já formatado: uma coluna que
não soma é a primeira coisa que faz alguém desconfiar da planilha inteira.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from julius.reporting.view_models import OpportunityVM, ReportViewModel

_HEADER_FILL = PatternFill("solid", fgColor="1F2933")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_MONEY = '"US$" #,##0.00'

#: Aba → de onde vêm as linhas. A ordem é a de execução, não a alfabética.
BUCKETS: tuple[tuple[str, str], ...] = (
    ("Fazer agora", "do_now"),
    ("Planejar", "plan"),
    ("Investigar primeiro", "investigate"),
    ("Monitorar", "monitor"),
)

#: Coluna → como extrair da `OpportunityVM`. Só leitura; nenhum cálculo.
COLUMNS: tuple[tuple[str, str, str], ...] = (
    ("ID", "id", "text"),
    ("Ativo", "asset", "text"),
    ("Serviço", "category_label", "text"),
    ("Achado", "title", "text"),
    ("Ação recomendada", "action", "text"),
    ("Economia/mês", "monthly_fmt", "money"),
    # As três colunas do lado da economia, e é onde elas têm de ficar: quem lê
    # "Indisponível" pergunta por quê na mesma linha, não três abas adiante.
    ("Por que sem cifra", "blocked_missing", "text"),
    ("Quem destrava", "blocked_unblocked_by", "text"),
    ("Como destravar", "blocked_next_action", "text"),
    ("Faixa", "band_fmt", "text"),
    ("Realizável no ano", "year_fmt", "money"),
    ("Baseline", "baseline_fmt", "money"),
    ("Qualidade da evidência", "evidence_quality_label", "text"),
    ("Ancorado na fatura", "anchored_in_billing", "bool"),
    ("Confiança", "confidence_label", "text"),
    ("Dificuldade", "difficulty_label", "text"),
    ("Prioridade", "exec_priority", "number"),
    ("Acionável", "actionable", "bool"),
    ("Próxima ação", "next_action", "text"),
    ("Dono", "owner", "text"),
    ("Processo", "source_process", "text"),
    ("Custo do processo/mês", "process_monthly_fmt", "money"),
    ("Evidência", "evidence", "lines"),
    ("Como aplicar", "how_to_apply", "text"),
    ("Como validar", "how_to_validate", "text"),
    ("Documentação", "doc_url", "text"),
)


def write_workbook(vm: ReportViewModel, path: str | Path) -> Path:
    """Escreve a planilha e devolve o caminho."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _summary_sheet(workbook.create_sheet("Resumo"), vm)
    for title, attribute in BUCKETS:
        rows = getattr(vm, attribute, [])
        _opportunities_sheet(workbook.create_sheet(title), rows)
    if vm.signals:
        _signals_sheet(workbook.create_sheet("Medições pendentes"), vm)
    _health_sheet(workbook.create_sheet("Saúde da coleta"), vm)
    _assumptions_sheet(workbook.create_sheet("Premissas"), vm)

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def _summary_sheet(sheet: Worksheet, vm: ReportViewModel) -> None:
    """O que o cabeçalho do HTML diz, em pares chave/valor."""
    _header(sheet, ["Indicador", "Valor"])
    rows = [
        ("Conta", vm.account_masked),
        ("Região", vm.region),
        ("Período", vm.period),
        ("Janela", vm.lookback),
        ("Scan", vm.scan_id),
        ("Gerado em", vm.generated_at),
        (vm.billing_period_label, vm.total_cost_fmt),
        ("Economia identificada/mês", vm.identified_fmt),
        ("Alta confiança/mês", vm.high_conf_fmt),
        ("Realizável no ano", vm.realizable_year_fmt),
        ("Oportunidades", vm.kpi_total),
        ("Acionáveis", vm.kpi_actionable),
        ("Estratégicas", vm.kpi_strategic),
        ("Saúde da coleta", vm.collection_status_label),
    ]
    for key, value in rows:
        sheet.append([key, value])
    _autosize(sheet, [34, 46])


def _opportunities_sheet(sheet: Worksheet, rows: list[OpportunityVM]) -> None:
    _header(sheet, [label for label, _, _ in COLUMNS])
    for item in rows:
        sheet.append([_cell(item, attribute, kind) for _, attribute, kind in COLUMNS])
    for index, (_, _, kind) in enumerate(COLUMNS, start=1):
        if kind != "money":
            continue
        letter = get_column_letter(index)
        for cell in sheet[letter][1:]:
            cell.number_format = _MONEY
    # Congelar o cabeçalho e ligar o filtro: a planilha existe para recortar.
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = sheet.dimensions
    _autosize(sheet, [18, 30, 14, 44, 44] + [16] * (len(COLUMNS) - 5))


def _health_sheet(sheet: Worksheet, vm: ReportViewModel) -> None:
    _header(
        sheet,
        [
            "Fonte",
            "Status",
            "Obrigatória",
            "Cobertura",
            "Origem",
            "Ações IAM ausentes",
            "Recursos afetados",
            "Categoria",
            "Impacto",
            "Próxima ação",
        ],
    )
    for item in vm.collection_health:
        sheet.append(
            [
                item.get("source", ""),
                item.get("status_label", item.get("status", "")),
                "sim" if item.get("required") else "não",
                item.get("coverage", ""),
                item.get("origin_label", ""),
                item.get("iam_actions", ""),
                item.get("iam_affected_resources", 0),
                item.get("error_category", ""),
                item.get("impact", ""),
                item.get("next_action", ""),
            ]
        )
    sheet.freeze_panes = "A2"
    _autosize(sheet, [28, 16, 12, 12, 24, 42, 18, 22, 52, 52])


def _assumptions_sheet(sheet: Worksheet, vm: ReportViewModel) -> None:
    """Onde a discussão sobre um número costuma terminar.

    Junta o manifesto do scan com as premissas de cada estimativa, porque
    "de onde veio esse valor" é a pergunta que a planilha precisa responder
    sem ninguém abrir o código.
    """
    _header(sheet, ["Escopo", "Item", "Detalhe"])
    for entry in vm.manifest:
        sheet.append(["Execução", entry.get("k", ""), entry.get("v", "")])

    seen: set[tuple[str, str]] = set()
    for bucket, attribute in BUCKETS:
        del bucket
        for item in getattr(vm, attribute, []):
            for assumption in getattr(item, "assumptions", ()) or ():
                key = (item.id, assumption)
                if key in seen:
                    continue
                seen.add(key)
                sheet.append([item.id, item.title, assumption])
    sheet.freeze_panes = "A2"
    _autosize(sheet, [24, 44, 80])


def _cell(item: OpportunityVM, attribute: str, kind: str) -> Any:
    value = getattr(item, attribute, None)
    if kind == "bool":
        return "sim" if value else "não"
    if kind == "lines":
        return "\n".join(str(line) for line in value or ())
    if kind == "money":
        return _as_number(value)
    if value is None:
        return ""
    return value


def _as_number(value: Any) -> Any:
    """`US$ 1.234,56` de volta a número, para a coluna somar.

    Texto que parece dinheiro e não soma é a primeira coisa que faz alguém
    desconfiar da planilha inteira. Quando não dá para converter — "Estratégico",
    "Indisponível" — o texto original fica, porque é informação.
    """
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip()
    if not text:
        return ""
    cleaned = (
        text.replace("US$", "").replace("R$", "").replace("\xa0", "").strip()
    )
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return text


def _signals_sheet(sheet: Worksheet, vm: ReportViewModel) -> None:
    """Medições pendentes: o que falta saber, quem descobre e quanto custa.

    A aba respondia "o que foi observado" e o leitor tinha de deduzir o que fazer.
    Agora a primeira coluna útil é a ação de remediação, a segunda é o próximo
    passo e a terceira diz quem o dá — porque três das quatro linhas costumam
    destravar com uma permissão IAM, não com meio dia de benchmark de alguém.

    A ordenação é a de `investigation_ranking_key`: retorno ÷ esforço de medir.

    A aba fica separada das oportunidades de propósito. Uma coluna "Economia"
    ao lado de uma faixa de ordem de grandeza convida a somar as duas, e somar
    hipótese com achado é como um relatório passa a prometer o que não mediu —
    por isso o cabeçalho diz "potencial" em todas as três colunas de valor.
    """
    _header(
        sheet,
        [
            "Ativo",
            "Ação de remediação",
            "Próximo passo",
            "Quem destrava",
            "Esforço de medir",
            "Fonte com lacuna",
            "Como destravar a fonte",
            "Regra",
            "Observação",
            "Potencial mínimo",
            "Potencial provável",
            "Potencial máximo",
            "O que a faixa assume",
        ],
    )
    por_regra = {
        (item["rule_id"], item["asset_name"]): item
        for item in vm.pending.get("items", [])
    }
    for signal in vm.signals:
        faixa = signal.get("potential_range") or {}
        pendente = por_regra.get(
            (signal.get("rule_id", ""), signal.get("asset_name", "")), {}
        )
        sheet.append(
            [
                signal.get("asset_name", ""),
                pendente.get("family_label", ""),
                signal.get("next_action", ""),
                pendente.get("unblocked_by", ""),
                signal.get("measurement_effort"),
                pendente.get("blocked_source", ""),
                pendente.get("source_next_action", ""),
                signal.get("rule_id", ""),
                signal.get("observation", ""),
                faixa.get("low"),
                faixa.get("expected"),
                faixa.get("high"),
                faixa.get("caveat", ""),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize(sheet, [26, 30, 44, 52, 16, 16, 16, 30, 44, 44])


def _header(sheet: Worksheet, labels: list[str]) -> None:
    sheet.append(labels)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28


def _autosize(sheet: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

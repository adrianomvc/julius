"""A planilha existe para filtrar e somar — não para reproduzir o HTML."""

from __future__ import annotations

from datetime import date

import pytest
from openpyxl import load_workbook

from julius.pipeline import analyze
from julius.reporting import excel
from julius.reporting.excel import BUCKETS, _as_number

SAMPLE = "data/sample/consumer-avi.json"


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    analysis = analyze(SAMPLE, today=date(2026, 7, 25), scan_id="xlsx")
    target = tmp_path_factory.mktemp("excel") / "report.xlsx"
    excel.write_workbook(analysis.vm, target)
    return load_workbook(target)


@pytest.fixture(scope="module")
def analysis():
    return analyze(SAMPLE, today=date(2026, 7, 25), scan_id="xlsx")


def test_there_is_a_sheet_per_bucket_plus_health_and_assumptions(workbook):
    # "Sinais" só existe quando há hipótese em aberto: uma aba vazia num
    # relatório entregue se lê como "procuramos e não achamos nada".
    assert [nome for nome in workbook.sheetnames if nome != "Sinais"] == [
        "Resumo",
        "Fazer agora",
        "Planejar",
        "Investigar primeiro",
        "Monitorar",
        "Saúde da coleta",
        "Premissas",
    ]


def test_every_opportunity_reaches_exactly_one_bucket_sheet(workbook, analysis):
    """A planilha não pode perder achado nem duplicar."""
    in_sheets = sum(
        workbook[title].max_row - 1 for title, _ in BUCKETS
    )
    in_report = sum(len(getattr(analysis.vm, attribute)) for _, attribute in BUCKETS)

    assert in_sheets == in_report
    assert in_sheets > 0

    ids = [
        workbook[title].cell(row, 1).value
        for title, _ in BUCKETS
        for row in range(2, workbook[title].max_row + 1)
    ]
    assert len(ids) == len(set(ids)), "achado apareceu em mais de uma aba"


def test_money_columns_are_numbers_not_text(workbook):
    """Coluna de dinheiro que não soma faz desconfiar da planilha inteira."""
    sheet = workbook["Monitorar"]
    header = [cell.value for cell in sheet[1]]
    column = header.index("Economia/mês") + 1

    values = [
        sheet.cell(row, column).value for row in range(2, sheet.max_row + 1)
    ]
    numeric = [value for value in values if isinstance(value, (int, float))]

    assert numeric, "nenhuma economia virou número"
    assert all(
        sheet.cell(row, column).number_format.startswith('"US$"')
        for row in range(2, sheet.max_row + 1)
        if isinstance(sheet.cell(row, column).value, (int, float))
    )


@pytest.mark.parametrize(
    "text,expected",
    [
        ("US$ 1.234,56", 1234.56),
        ("US$ 114,27", 114.27),
        ("US$ 0,42", 0.42),
        (2.5, 2.5),
        ("", ""),
    ],
)
def test_formatted_money_becomes_a_number_again(text, expected):
    assert _as_number(text) == expected


@pytest.mark.parametrize("text", ["Estratégico", "Indisponível", "—"])
def test_what_is_not_a_number_keeps_its_text(text):
    """'Estratégico' é informação; virar zero seria mentira."""
    assert _as_number(text) == text


def test_the_bucket_sheets_are_filterable(workbook):
    for title, _ in BUCKETS:
        sheet = workbook[title]
        assert sheet.freeze_panes == "A2"
        if sheet.max_row > 1:
            assert sheet.auto_filter.ref, f"{title} sem filtro"


def test_the_assumptions_sheet_explains_where_numbers_came_from(workbook):
    sheet = workbook["Premissas"]
    rows = [
        [cell.value for cell in sheet[row]] for row in range(2, sheet.max_row + 1)
    ]

    assert rows, "aba de premissas vazia — é onde a discussão termina"
    scopes = {row[0] for row in rows}
    assert "Execução" in scopes, "o manifesto do scan precisa estar lá"
    # E as premissas de cálculo de cada achado, não só o manifesto.
    assert len(scopes) > 1
    # A procedência do número, seja qual for: tarifa versionada quando o valor
    # é modelado, cobrança rateada quando ele está ancorado na fatura.
    procedencia = ("tarifa", "USD", "cobrança", "ratead", "faixa da regra")
    assert any(
        any(marca in str(row[2]) for marca in procedencia) for row in rows
    ), "a aba precisa dizer de onde cada número veio"


def test_the_sheet_reports_the_evidence_quality_scale(workbook):
    sheet = workbook["Monitorar"]
    header = [cell.value for cell in sheet[1]]

    assert "Qualidade da evidência" in header
    assert "Ancorado na fatura" in header
    column = header.index("Ancorado na fatura") + 1
    values = {
        sheet.cell(row, column).value for row in range(2, sheet.max_row + 1)
    }
    assert values <= {"sim", "não"}


def test_the_workbook_computes_nothing_of_its_own(analysis, tmp_path):
    """A planilha lê a mesma view model do HTML; divergir seria bug em dobro."""
    target = excel.write_workbook(analysis.vm, tmp_path / "r.xlsx")
    sheet = load_workbook(target)["Resumo"]
    summary = {
        sheet.cell(row, 1).value: sheet.cell(row, 2).value
        for row in range(2, sheet.max_row + 1)
    }

    assert summary["Economia identificada/mês"] == analysis.vm.identified_fmt
    assert summary["Realizável no ano"] == analysis.vm.realizable_year_fmt
    assert summary["Oportunidades"] == analysis.vm.kpi_total
    assert summary["Saúde da coleta"] == analysis.vm.collection_status_label
